from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
from datetime import datetime

print("sample test case started")

# Replace 'path/to/your/excel/file.xlsx' with the actual path to your local Excel file
excel_file_path = 'C:/Users/ashra/PycharmProjects/AutomationOfSearchAndStore/sample.xlsx'

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file_path)

# Get the current weekday (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
current_weekday = datetime.today().weekday()

# Define a list of weekdays corresponding to sheet names
weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# Get the sheet name for the current weekday
current_sheet_name = weekday_names[current_weekday]

try:
    print(f"\nProcessing sheet: {current_sheet_name}")

    # Get the sheet
    sheet = workbook[current_sheet_name]

    # Assume the search inputs are in column "C" from rows 3 to 12
    search_inputs = [cell.value for cell in sheet['C'][2:12]]

    # Continue with the rest of your code...
    driver = webdriver.Chrome()
    driver.maximize_window()

    # Iterate through search inputs from the sheet
    for i, search_input in enumerate(search_inputs):

        # navigate to the URL
        driver.get("https://www.google.com/")

        # Wait for the search box to be present
        search_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.NAME, "q")))

        # Clear and enter the search input
        search_box.clear()
        search_box.send_keys(search_input)

        # Wait for the search suggestions to be visible
        wait = WebDriverWait(driver, 5)

        # Define a list of suggestion IDs
        suggestion_ids = ['jZ2SBf', 'ERWdKc', 'c7mM1c', 'Zrbbw', 'sBQTL', 'vTtioc', 'gLuoSb', 'mdzGzd', 'hwdq1',
                          'HFKcI']

        # Initialize lists to store suggestions
        shortest_suggestions = []
        longest_suggestions = []

        # Iterate over the suggestion IDs and capture suggestions
        for suggestion_id in suggestion_ids:
            locator = (By.CSS_SELECTOR, f"div[id='{suggestion_id}'] div:nth-child(1) span:nth-child(1)")
            suggestions = wait.until(EC.visibility_of_all_elements_located(locator))
            suggestions_text = [suggestion.text for suggestion in suggestions]
            shortest_suggestions.append(min(suggestions_text, key=len))
            longest_suggestions.append(max(suggestions_text, key=len))

        # Write the suggestions to the Excel sheet (assuming you want to write them in columns D and E)
        sheet.cell(row=i + 3, column=4, value=min(shortest_suggestions, key=len))
        sheet.cell(row=i + 3, column=5, value=max(longest_suggestions, key=len))

        # Print the suggestions for the current search input
        print(f"For input '{search_input}':")
        print(f"Shortest suggestion: {min(shortest_suggestions, key=len)}")
        print(f"Longest suggestion: {max(longest_suggestions, key=len)}")
        print('________________________________________________________________________')

        # Clear the lists for the next iteration
        shortest_suggestions.clear()
        longest_suggestions.clear()

        # Introduce a delay (adjust as needed)
        time.sleep(1)

finally:
    # Save the workbook
    workbook.save(excel_file_path)
    print("sample test case successfully completed")
